import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill

class Formater:

    @staticmethod
    def doFormat(file_path:str):

        """
        对 Excel 文件进行样式处理
        """

        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path)

        # 循环遍历每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet['A1']='方块名称'

            # 删除空白sheet
            if sheet.max_row==1:
                workbook.remove(sheet)

            # 设置单元格背景色
            for i in range(1,sheet.max_row+1):
                sheet.cell(row=i,column=1).fill=PatternFill(fill_type='solid',fgColor='b7ffb7')
            for i in range(1,sheet.max_column+1):
                sheet.cell(row=1,column=i).fill=PatternFill(fill_type='solid',fgColor='b7ffb7')

            # 设置格式，字体，字体颜色，字体大小，边框，居中显示，加粗
            for i in range(1,sheet.max_row+1):
                for j in range(1,sheet.max_column+1):
                    sheet.cell(row=i,column=j).alignment=Alignment(horizontal='center',
                                                                vertical='center',wrapText=False)
                    sheet.cell(row=i,column=j).font=Font(bold=True,name='宋体')
                    sheet.cell(row=i,column=j).border=Border(left=Side(border_style='thin',color='000000'),
                                                            right=Side(border_style='thin',color='000000'),
                                                            top=Side(border_style='thin',color='000000'),
                                                            bottom=Side(border_style='thin',color='000000'))

            for i in range(2,sheet.max_row+1):
                for j in range(2,sheet.max_column+1):
                    sheet.cell(row=i,column=j).font=Font(bold=False,name='Times New Roman')    

            # 循环遍历所有列，自适应列宽
            lks=[]
            for i in range(1,sheet.max_column+1):
                lk=1
                for j in range(1,sheet.max_row+1):
                    sz=sheet.cell(row=j,column=i).value
                    if isinstance(sz,str):
                        lk1=len(sz.encode('gbk'))
                    else:
                        lk1=len(str(sz))
                    if lk<lk1:
                        lk=lk1
                lks.append(lk)

            for i in range(1,sheet.max_column+1):
                k=get_column_letter(i)
                sheet.column_dimensions[k].width=lks[i-1]+2
        
        # 保存修改后的工作簿
        workbook.save(file_path)

        # 关闭工作簿
        workbook.close()
        